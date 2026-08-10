"""
Prvotkář 3.2 – FastAPI backend
Opravy: asyncio import, cache VR, IČO vyhledávání, ping endpoint
"""
from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import sqlite3, io, os, asyncio as _asyncio, time
import httpx
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import uvicorn

app = FastAPI(title="Prvotkář 3.2 API")

app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True,
                   allow_methods=["*"], allow_headers=["*"])

@app.get("/")
async def root():
    return FileResponse("index.html")

DB_FILE = "prvotkar.db"

def _migrate_db():
    """Doplní sloupec okres do starší databáze (jednorázově)."""
    if os.path.exists(DB_FILE):
        conn = sqlite3.connect(DB_FILE)
        try:
            cols = {r[1] for r in conn.execute("PRAGMA table_info(subjekty)")}
            if "okres" not in cols:
                conn.execute("ALTER TABLE subjekty ADD COLUMN okres TEXT")
                conn.execute("CREATE INDEX IF NOT EXISTS idx_okres ON subjekty(okres)")
                conn.commit()
        finally:
            conn.close()
_migrate_db()

# ── Cache pro VR osoby (24h) ──────────────────────────────────────────────────
_vr_cache: dict = {}   # ico -> (timestamp, data)
VR_CACHE_TTL = 86400   # 24 hodin

def get_db():
    if not os.path.exists(DB_FILE):
        raise HTTPException(
            status_code=503,
            detail="Databáze nenalezena. Spusť nejdřív: python3 sync_ares.py"
        )
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def row_to_dict(row):
    return dict(row)

def obec_filter(obec: str):
    """Vrátí SQL fragment a params pro hledání obce (přesná shoda + prefix-)."""
    return "(obec = ? OR obec LIKE ? OR obec LIKE ?)", [obec, obec + "-%", obec + " %"]

# ===================== ENDPOINTS =====================

@app.get("/api/ping")
async def ping():
    """Keep-alive endpoint pro UptimeRobot / monitoring."""
    return {"ok": True, "ts": int(time.time())}

@app.get("/api/version")
async def version():
    return {"version": "3.2", "ok": True}

@app.get("/api/kraje")
async def get_kraje():
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT DISTINCT kraj, kraj_kod FROM subjekty WHERE kraj IS NOT NULL ORDER BY kraj"
        ).fetchall()
        return [{"nazev": r["kraj"], "kod": r["kraj_kod"]} for r in rows]
    finally:
        conn.close()

@app.get("/api/okresy")
async def get_okresy(kraj: Optional[str] = None):
    """Seznam okresů (naplní se po novém běhu sync_ares.py)."""
    conn = get_db()
    try:
        q = "SELECT okres, COUNT(*) n FROM subjekty WHERE okres IS NOT NULL"
        params = []
        if kraj:
            q += " AND kraj = ?"
            params.append(kraj)
        q += " GROUP BY okres ORDER BY okres"
        return [{"okres": r[0], "pocet": r[1]} for r in conn.execute(q, params)]
    finally:
        conn.close()

@app.get("/api/svj")
async def get_svj(
    obec: Optional[str] = Query(None),
    okres: Optional[str] = None,
    ulice: Optional[str] = None,
    cast_obce: Optional[str] = None,
    typ: Optional[str] = "svj",
    start: int = 0,
    pocet: int = 2000,
):
    if not obec and not okres:
        raise HTTPException(status_code=400, detail="Zadejte obec nebo okres.")
    conn = get_db()
    try:
        if obec:
            sql_frag, params = obec_filter(obec)
        else:
            sql_frag, params = "1=1", []
        params.append(typ)
        filters = f"{sql_frag} AND typ = ?"
        if okres:
            filters += " AND okres = ?"
            params.append(okres)
        if cast_obce:
            filters += " AND cast_obce = ?"
            params.append(cast_obce)
        if ulice:
            filters += " AND ulice = ?"
            params.append(ulice)
        all_rows = conn.execute(
            f"SELECT * FROM subjekty WHERE {filters} ORDER BY nazev",
            params
        ).fetchall()
        celkem = len(all_rows)
        page = all_rows[start:start + pocet]

        subjekty = []
        for r in page:
            d = row_to_dict(r)
            subjekty.append({
                "ico": d["ico"],
                "obchodniJmeno": d["nazev"],
                "stavSubjektu": d["stav"],
                "datumVzniku": d["datum_vzniku"],
                "sidlo": {
                    "nazevObce": d["obec"],
                    "nazevCastiObce": d["cast_obce"],
                    "nazevUlice": d["ulice"],
                    "cisloDomovni": d["cislo_popisne"],
                    "cisloOrientacni": d["cislo_orientacni"],
                    "psc": d["psc"],
                    "nazevKraje": d["kraj"],
                    "nazevOkresu": d.get("okres"),
                },
                "lat": d.get("lat"),
                "lng": d.get("lng"),
            })
        return {"celkem": celkem, "subjekty": subjekty}
    finally:
        conn.close()

@app.get("/api/obce")
async def get_obce(q: str = Query(..., min_length=2), typ: Optional[str] = None):
    """Autocomplete obcí z lokální DB včetně počtu SVJ/BD."""
    conn = get_db()
    try:
        query = """SELECT obec, kraj, COUNT(*) as pocet
                   FROM subjekty WHERE obec LIKE ? AND obec IS NOT NULL"""
        params = [f"{q}%"]
        if typ:
            query += " AND typ = ?"
            params.append(typ)
        query += " GROUP BY obec ORDER BY pocet DESC, obec LIMIT 20"
        rows = conn.execute(query, params).fetchall()
        return [{"obec": r["obec"], "kraj": r["kraj"], "pocet": r["pocet"]} for r in rows]
    finally:
        conn.close()

@app.get("/api/ico/{ico}")
async def get_by_ico(ico: str):
    """Vyhledání přímo podle IČO – vrátí základní data bez detailu."""
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM subjekty WHERE ico = ?", [ico]).fetchone()
    finally:
        conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="IČO nenalezeno")
    d = row_to_dict(row)
    return {
        "ico": d["ico"],
        "obchodniJmeno": d["nazev"],
        "stavSubjektu": d["stav"],
        "datumVzniku": d["datum_vzniku"],
        "sidlo": {
            "nazevObce": d["obec"],
            "nazevCastiObce": d["cast_obce"],
            "nazevUlice": d["ulice"],
            "cisloDomovni": d["cislo_popisne"],
            "cisloOrientacni": d["cislo_orientacni"],
            "psc": d["psc"],
            "nazevKraje": d["kraj"],
        }
    }

@app.get("/api/stats")
async def get_stats():
    conn = get_db()
    try:
        total   = conn.execute("SELECT COUNT(*) FROM subjekty").fetchone()[0]
        svj     = conn.execute("SELECT COUNT(*) FROM subjekty WHERE typ='svj'").fetchone()[0]
        bd      = conn.execute("SELECT COUNT(*) FROM subjekty WHERE typ='bd'").fetchone()[0]
        updated = conn.execute("SELECT MAX(updated_at) FROM subjekty").fetchone()[0]
        obce    = conn.execute("SELECT COUNT(DISTINCT obec) FROM subjekty").fetchone()[0]
        return {"celkem": total, "svj": svj, "bd": bd, "obce": obce, "posledni_sync": updated}
    finally:
        conn.close()

@app.get("/api/hledat")
async def hledat(q: str = Query(..., min_length=2), typ: Optional[str] = "svj", limit: int = 30):
    """Fulltext vyhledávání v celé DB podle názvu."""
    conn = get_db()
    try:
        rows = conn.execute(
            """SELECT * FROM subjekty
               WHERE (nazev LIKE ? OR ico LIKE ?) AND typ = ?
               ORDER BY nazev LIMIT ?""",
            [f"%{q}%", f"%{q}%", typ, limit]
        ).fetchall()
        subjekty = []
        for r in rows:
            d = row_to_dict(r)
            subjekty.append({
                "ico": d["ico"],
                "obchodniJmeno": d["nazev"],
                "stavSubjektu": d["stav"],
                "sidlo": {
                    "nazevObce": d["obec"],
                    "nazevUlice": d["ulice"],
                    "cisloDomovni": d["cislo_popisne"],
                    "psc": d["psc"],
                }
            })
        return {"subjekty": subjekty, "celkem": len(subjekty)}
    finally:
        conn.close()

def _parse_osoby_vr(vr_json: dict) -> list:
    osoby = []
    seen  = set()
    for zaznam in vr_json.get("zaznamy", []):
        for org in zaznam.get("statutarniOrgany", []):
            for clen in org.get("clenoveOrganu", []):
                if clen.get("datumVymazu"):
                    continue
                clenstvi    = clen.get("clenstvi", {})
                funkce_info = clenstvi.get("funkce", {})
                if funkce_info.get("zanikFunkce"):
                    continue
                fo = clen.get("fyzickaOsoba")
                if not fo:
                    po = clen.get("pravnickaOsoba", {})
                    for z in po.get("zastoupeni", []):
                        if not z.get("datumVymazu"):
                            fo = z.get("fyzickaOsoba")
                            break
                if not fo:
                    continue
                jmeno = " ".join(filter(None, [
                    fo.get("titulPredJmenem", ""),
                    fo.get("jmeno", ""),
                    fo.get("prijmeni", "")
                ])).strip()
                if fo.get("titulZaJmenem"):
                    jmeno += ", " + fo["titulZaJmenem"]
                narozeni = fo.get("datumNarozeni", "")
                key = f"{jmeno}|{narozeni}"
                if not jmeno or key in seen:
                    continue
                seen.add(key)
                osoby.append({
                    "jmeno":         jmeno,
                    "prijmeni":      fo.get("prijmeni", ""),
                    "datumNarozeni": narozeni,
                    "funkce":        funkce_info.get("nazev", "člen výboru"),
                    "nazevRole":     funkce_info.get("nazev", "člen výboru"),
                })
    return osoby

@app.get("/api/svj/{ico}/detail")
async def get_svj_detail(ico: str):
    import re as _re
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM subjekty WHERE ico = ?", [ico]).fetchone()
    finally:
        conn.close()

    if not row:
        raise HTTPException(status_code=404, detail="Subjekt nenalezen")

    d    = row_to_dict(row)
    base = {
        "ico":          d["ico"],
        "nazev":        d["nazev"],
        "sidlo": {
            "nazevObce":        d["obec"],
            "nazevCastiObce":   d["cast_obce"],
            "nazevUlice":       d["ulice"],
            "cisloDomovni":     d["cislo_popisne"],
            "cisloOrientacni":  d["cislo_orientacni"],
            "psc":              d["psc"],
            "nazevKraje":       d["kraj"],
        },
        "datumVzniku":  d["datum_vzniku"],
        "stavSubjektu": d["stav"],
        "osoby":        [],
        "spisovaZnacka": None,
        "subjektId":    None,
    }

    # ── Cache VR osoby 24h ──
    cached = _vr_cache.get(ico)
    if cached and (time.time() - cached[0]) < VR_CACHE_TTL:
        base["osoby"]        = cached[1].get("osoby", [])
        base["spisovaZnacka"] = cached[1].get("spisovaZnacka")
        base["subjektId"]    = cached[1].get("subjektId")
        return base

    try:
        async with httpx.AsyncClient(timeout=15) as client:
            r = await client.get(
                f"https://ares.gov.cz/ekonomicke-subjekty-v-be/rest/ekonomicke-subjekty-vr/{ico}"
            )
            if r.status_code == 200:
                vr = r.json()
                osoby = _parse_osoby_vr(vr)
                for zaznam in vr.get("zaznamy", []):
                    for sz in zaznam.get("spisovaZnacka", []):
                        oddil  = sz.get("oddil", "")
                        vlozka = sz.get("vlozka", "")
                        soud   = sz.get("soud", "")
                        if oddil and vlozka:
                            base["spisovaZnacka"] = f"{oddil} {vlozka}/{soud}"
                base["osoby"] = osoby

            r2 = await client.get(
                f"https://or.justice.cz/ias/ui/rejstrik-$firma?ico={ico}&jenPlatne=PLATNE",
                headers={"User-Agent": "Mozilla/5.0"}, timeout=8
            )
            if r2.status_code == 200:
                ids = _re.findall(r'subjektId[=:](\d+)', r2.text)
                if ids:
                    base["subjektId"] = ids[0]

        # Ulož do cache
        _vr_cache[ico] = (time.time(), {
            "osoby":        base["osoby"],
            "spisovaZnacka": base["spisovaZnacka"],
            "subjektId":    base["subjektId"],
        })
    except Exception:
        pass

    return base

@app.get("/api/casti")
async def get_casti(obec: str = Query(...), typ: Optional[str] = None):
    conn = get_db()
    try:
        sql_frag, params = obec_filter(obec)
        q = f"SELECT DISTINCT cast_obce FROM subjekty WHERE {sql_frag} AND cast_obce IS NOT NULL AND cast_obce != obec"
        if typ:
            q += " AND typ = ?"
            params.append(typ)
        q += " ORDER BY cast_obce"
        rows = conn.execute(q, params).fetchall()
        return [r[0] for r in rows]
    finally:
        conn.close()

@app.get("/api/ulice")
async def get_ulice(obec: str = Query(...), cast_obce: Optional[str] = None, typ: Optional[str] = None):
    conn = get_db()
    try:
        sql_frag, params = obec_filter(obec)
        q = f"SELECT DISTINCT ulice FROM subjekty WHERE {sql_frag} AND ulice IS NOT NULL"
        if cast_obce:
            q += " AND cast_obce = ?"
            params.append(cast_obce)
        if typ:
            q += " AND typ = ?"
            params.append(typ)
        q += " ORDER BY ulice"
        rows = conn.execute(q, params).fetchall()
        return [r[0] for r in rows]
    finally:
        conn.close()

@app.get("/api/export/excel")
async def export_excel(
    obec: str = Query(...),
    ulice: Optional[str] = None,
    cast_obce: Optional[str] = None,
    typ: Optional[str] = "svj",
):
    result = await get_svj(obec=obec, ulice=ulice, cast_obce=cast_obce, typ=typ, start=0, pocet=9999)
    data   = result["subjekty"]
    label  = "BD" if typ == "bd" else "SVJ"

    MAX_OSOBY = 150
    ico_list  = [s.get("ico") for s in data[:MAX_OSOBY] if s.get("ico")]
    osoby_map: dict = {}

    async with httpx.AsyncClient(timeout=12) as client:
        sem = _asyncio.Semaphore(6)
        async def fetch_osoby(ico):
            async with sem:
                # Check cache first
                cached = _vr_cache.get(ico)
                if cached and (time.time() - cached[0]) < VR_CACHE_TTL:
                    osoby_map[ico] = cached[1].get("osoby", [])
                    return
                try:
                    await _asyncio.sleep(0.1)
                    r = await client.get(
                        f"https://ares.gov.cz/ekonomicke-subjekty-v-be/rest/ekonomicke-subjekty-vr/{ico}"
                    )
                    if r.status_code == 200:
                        osoby_map[ico] = _parse_osoby_vr(r.json())
                except Exception:
                    pass
        await _asyncio.gather(*[fetch_osoby(ico) for ico in ico_list])

    wb = Workbook()
    ws = wb.active
    ws.title = "Seznam"

    hf  = PatternFill(start_color="8B0C1E", end_color="8B0C1E", fill_type="solid")
    hf2 = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    af  = PatternFill(start_color="FDF0F2", end_color="FDF0F2", fill_type="solid")
    b   = Border(
        left=Side(style="thin", color="BDC3C7"), right=Side(style="thin", color="BDC3C7"),
        top=Side(style="thin", color="BDC3C7"),  bottom=Side(style="thin", color="BDC3C7")
    )

    max_ostatni = max(
        (len([o for o in v if "předseda" not in (o.get("funkce","") or "").lower()])
         for v in osoby_map.values()), default=0
    )
    base_headers = [
        "IČO", f"Název {label}", "Ulice", "ČP/CO", "Obec", "PSČ", "Kraj", "Rok vzniku",
        "Předseda – jméno", "Předseda – narozen",
        "Místopředseda – jméno", "Místopředseda – narozen"
    ]
    total_cols = len(base_headers) + max_ostatni * 3

    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    ws["A1"].value = f"Seznam {label} – {obec}" + (f" / {ulice}" if ulice else "")
    ws["A1"].font  = Font(bold=True, size=14, name="Calibri", color="8B0C1E")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{get_column_letter(total_cols)}2")
    ws["A2"].value = f"Celkem: {len(data)} záznamů | Zdroj: ARES | Výbor: živá data z ARES VR | IP Polná"
    ws["A2"].font  = Font(italic=True, size=10, color="7F8C8D", name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="center")

    for col, header in enumerate(base_headers, 1):
        c = ws.cell(row=4, column=col, value=header)
        c.fill = hf if col <= 8 else hf2
        c.font = Font(color="FFFFFF", bold=True, size=10 if col > 8 else 11, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = b

    for i in range(max_ostatni):
        cb = len(base_headers) + 1 + i * 3
        for offset, header in enumerate([f"Člen {i+1} – jméno", f"Člen {i+1} – funkce", f"Člen {i+1} – narozen"]):
            c = ws.cell(row=4, column=cb + offset, value=header)
            c.fill = hf2
            c.font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = b
    ws.row_dimensions[4].height = 30

    for ri, s in enumerate(data, 5):
        sidlo = s.get("sidlo") or {}
        cp    = str(sidlo.get("cisloDomovni") or "")
        co    = sidlo.get("cisloOrientacni") or ""
        cislo = cp + ("/" + str(co) if co else "")
        rok   = (s.get("datumVzniku") or "")[:4]
        fill  = af if (ri - 5) % 2 == 1 else None

        base_vals = [
            s.get("ico"), s.get("obchodniJmeno"),
            sidlo.get("nazevUlice") or "", cislo,
            sidlo.get("nazevObce"), sidlo.get("psc"),
            sidlo.get("nazevKraje"), rok
        ]
        for col, val in enumerate(base_vals, 1):
            c = ws.cell(row=ri, column=col, value=val)
            if fill: c.fill = fill
            c.border = b
            c.font = Font(size=10, name="Calibri")
            c.alignment = Alignment(vertical="center")

        osoby   = osoby_map.get(s.get("ico"), [])
        predseda = next((o for o in osoby if "předseda" in (o.get("funkce","")).lower()
                         and "místopředs" not in (o.get("funkce","")).lower()), {})
        mistoprds = next((o for o in osoby if "místopředs" in (o.get("funkce","")).lower()), {})
        ostatni   = [o for o in osoby if o is not predseda and o is not mistoprds]

        def put(row, col, val):
            c = ws.cell(row=row, column=col, value=val)
            c.border = b
            c.font = Font(size=10, name="Calibri")
            c.alignment = Alignment(vertical="center")
            if fill: c.fill = fill

        put(ri, 9,  predseda.get("jmeno", ""))
        put(ri, 10, predseda.get("datumNarozeni", "") or predseda.get("narozeni", ""))
        put(ri, 11, mistoprds.get("jmeno", ""))
        put(ri, 12, mistoprds.get("datumNarozeni", "") or mistoprds.get("narozeni", ""))
        for i, osoba in enumerate(ostatni):
            cb = 13 + i * 3
            put(ri, cb,   osoba.get("jmeno", ""))
            put(ri, cb+1, osoba.get("funkce", ""))
            put(ri, cb+2, osoba.get("datumNarozeni", "") or osoba.get("narozeni", ""))

    for i, w in enumerate([12, 45, 25, 14, 20, 9, 22, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(total_cols)}{4 + len(data)}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fn = f"{label}_{obec.replace(' ', '_')}.xlsx"
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"})

# ===================== SYNC =====================
_sync_running = False
_sync_status  = {"running": False, "progress": "", "done": False, "error": "", "pct": 0, "eta": ""}

@app.get("/api/sync/status")
async def sync_status():
    conn = get_db()
    try:
        svj     = conn.execute("SELECT COUNT(*) FROM subjekty WHERE typ='svj'").fetchone()[0]
        bd      = conn.execute("SELECT COUNT(*) FROM subjekty WHERE typ='bd'").fetchone()[0]
        updated = conn.execute("SELECT MAX(updated_at) FROM subjekty").fetchone()[0]
    finally:
        conn.close()
    return {**_sync_status, "svj": svj, "bd": bd, "posledni_sync": updated}

@app.post("/api/sync/start")
async def sync_start():
    global _sync_running, _sync_status
    if _sync_running:
        return {"ok": False, "msg": "Sync už běží"}
    _sync_running = True
    _sync_status  = {"running": True, "progress": "Spouštím sync…", "done": False, "error": "", "pct": 0, "eta": ""}
    _asyncio.create_task(_run_sync())
    return {"ok": True, "msg": "Sync spuštěn"}

async def _run_sync():
    import re as _re
    global _sync_running, _sync_status
    try:
        proc = await _asyncio.create_subprocess_exec(
            __import__("sys").executable, "sync_ares.py",
            stdout=_asyncio.subprocess.PIPE,
            stderr=_asyncio.subprocess.STDOUT,
            cwd=os.path.dirname(os.path.abspath(__file__))
        )
        last = ""
        while True:
            line = await proc.stdout.readline()
            if not line:
                break
            last = line.decode("utf-8", errors="replace").strip()
            if last:
                _sync_status["progress"] = last
                m = _re.search(r'\[\s*(\d+\.?\d*)%\]', last)
                if m:
                    _sync_status["pct"] = float(m.group(1))
                m2 = _re.search(r'~([\dhms ]+)zbyvá', last)
                if m2:
                    _sync_status["eta"] = m2.group(1).strip()
        await proc.wait()
        _sync_status = {"running": False, "progress": last, "done": True, "error": "", "pct": 100, "eta": ""}
    except Exception as e:
        _sync_status = {"running": False, "progress": "", "done": False, "error": str(e), "pct": 0, "eta": ""}
    finally:
        _sync_running = False

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=False)
