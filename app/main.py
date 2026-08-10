import hashlib
import os
import shutil
import tempfile
import threading
import time
from datetime import datetime
from pathlib import Path

from fastapi import (FastAPI, Depends, HTTPException, UploadFile, File, Form,
                     Header)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse
from pydantic import BaseModel, Field
from sqlalchemy import select, desc, func
from sqlalchemy.orm import Session

from .db import init_db, get_db, SessionLocal
from .models import Subject, Document, Signal
from .signal_engine import lead_level
from .pipeline import ingest_text, ingest_pdf, sync_many, SYNC_STATE
from .import_justice import import_dataset

app = FastAPI(title="RBD Radar", version="0.3.0")

# CORS: umožňuje frontendu Prvotkáře (nebo jiné aplikaci) číst API Radaru.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["GET", "POST"],
    allow_headers=["*"],
)

STATIC_DIR = Path(__file__).parent / "static"


# ---------------------------------------------------------------------------
# Synchronizace na pozadí + denní plánovač
# ---------------------------------------------------------------------------

def _notify_hot(results: list[dict]):
    """Pošle e-mail s novými HOT/HIGH leady (pokud je nastaveno SMTP).

    Potřebné proměnné prostředí: RADAR_SMTP_HOST, RADAR_SMTP_USER,
    RADAR_SMTP_PASS, RADAR_NOTIFY_TO (volitelně RADAR_SMTP_PORT, vých. 587).
    Bez nich se notifikace tiše přeskočí.
    """
    host = os.getenv("RADAR_SMTP_HOST")
    to = os.getenv("RADAR_NOTIFY_TO")
    if not host or not to:
        return

    hot = []
    for r in results:
        for d in r.get("documents", []):
            if (not d.get("duplicate") and not d.get("error")
                    and d.get("score", 0) >= 60):
                hot.append((r.get("name", r.get("ico", "?")), r.get("ico"),
                            d.get("score"), d.get("lead_level"),
                            [s.get("label") + (f" = {s['value']}" if s.get("value") else "")
                             for s in d.get("signals", [])[:6]]))
    if not hot:
        return

    try:
        import smtplib
        from email.message import EmailMessage
        msg = EmailMessage()
        msg["Subject"] = f"RBD Radar: {len(hot)} nových nadějných leadů 🔥"
        msg["From"] = os.getenv("RADAR_SMTP_USER", "radar@localhost")
        msg["To"] = to
        lines = []
        for name, ico, score, level, signals in hot:
            lines.append(f"• {name} (IČO {ico}) — {score}/100 {level}")
            for s in signals:
                lines.append(f"    · {s}")
        lines.append("")
        lines.append("Dashboard: https://rbd-radar.onrender.com")
        msg.set_content("\n".join(lines))
        with smtplib.SMTP(host, int(os.getenv("RADAR_SMTP_PORT", "587")),
                          timeout=30) as smtp:
            smtp.starttls()
            user = os.getenv("RADAR_SMTP_USER")
            if user:
                smtp.login(user, os.getenv("RADAR_SMTP_PASS", ""))
            smtp.send_message(msg)
        print(f"Notifikace odeslána: {len(hot)} leadů -> {to}")
    except Exception as exc:
        print(f"Notifikace se nepodařila: {exc}")


def _run_sync_job(limit: int, city: str | None, max_docs: int,
                  since_days: int | None, icos: list[str] | None = None):
    SYNC_STATE.update(running=True, started_at=datetime.utcnow().isoformat(),
                      finished_at=None, progress="startuji…",
                      processed_subjects=0, new_documents=0, hot_found=0,
                      error=None)
    db = SessionLocal()
    try:
        results = sync_many(db, limit=limit, city=city, max_docs=max_docs,
                            since_days=since_days, state=SYNC_STATE, icos=icos)
        SYNC_STATE["progress"] = "hotovo"
        _notify_hot(results)
    except Exception as exc:
        SYNC_STATE["error"] = str(exc)
        SYNC_STATE["progress"] = "chyba"
    finally:
        db.close()
        SYNC_STATE["running"] = False
        SYNC_STATE["finished_at"] = datetime.utcnow().isoformat()


def _start_sync(limit: int, city: str | None, max_docs: int,
                since_days: int | None, icos: list[str] | None = None) -> bool:
    if SYNC_STATE["running"]:
        return False
    threading.Thread(
        target=_run_sync_job,
        args=(limit, city, max_docs, since_days, icos),
        daemon=True,
    ).start()
    return True


def _scheduler():
    """Automatické synchronizace uvnitř webové služby (bez placeného cronu).

    RADAR_DAILY_SYNC=1  — ranní kontrola novinek (RADAR_SYNC_HOUR_UTC, vých. 5;
                          RADAR_SYNC_LIMIT, vých. 15 SVJ; jen čerstvé listiny)
    RADAR_NIGHT_SYNC=1  — noční dlouhý běh (RADAR_NIGHT_HOUR_UTC, vých. 0;
                          RADAR_NIGHT_LIMIT, vých. 150 SVJ; hlubší stahování)
    """
    daily = os.getenv("RADAR_DAILY_SYNC") == "1"
    night = os.getenv("RADAR_NIGHT_SYNC") == "1"
    daily_hour = int(os.getenv("RADAR_SYNC_HOUR_UTC", "5"))
    night_hour = int(os.getenv("RADAR_NIGHT_HOUR_UTC", "0"))
    daily_limit = int(os.getenv("RADAR_SYNC_LIMIT", "15"))
    night_limit = int(os.getenv("RADAR_NIGHT_LIMIT", "150"))
    last_daily = last_night = None
    while True:
        now = datetime.utcnow()
        if night and now.hour == night_hour and last_night != now.date():
            last_night = now.date()
            # Dlouhý běh: víc SVJ i listin na subjekt; poběží klidně hodiny,
            # klient drží pauzy, aby nedráždil justice.cz.
            _start_sync(limit=night_limit, city=None, max_docs=5,
                        since_days=None)
        elif daily and now.hour == daily_hour and last_daily != now.date():
            last_daily = now.date()
            _start_sync(limit=daily_limit, city=None, max_docs=3,
                        since_days=90)
        time.sleep(300)


@app.on_event("startup")
def startup():
    Path("data").mkdir(exist_ok=True)
    init_db()
    if os.getenv("RADAR_DAILY_SYNC") == "1" or os.getenv("RADAR_NIGHT_SYNC") == "1":
        threading.Thread(target=_scheduler, daemon=True).start()


class SubjectIn(BaseModel):
    ico: str = Field(min_length=6, max_length=20)
    name: str
    legal_form: str | None = None
    address: str | None = None
    source_url: str | None = None


class DocumentIn(BaseModel):
    ico: str
    external_id: str
    title: str
    source_url: str | None = None
    document_date: datetime | None = None
    text: str


class JusticeImportIn(BaseModel):
    dataset: str = "svj-actual-brno-2026"
    limit: int | None = None


class SyncIn(BaseModel):
    max_docs: int = 5


def require_api_key(x_api_key: str | None = Header(None)):
    """Ochrana zápisových endpointů při veřejném nasazení.

    Lokálně (bez nastaveného RADAR_API_KEY) se nic nevyžaduje.
    Na Renderu se klíč vygeneruje automaticky (viz render.yaml) a posílá se
    v hlavičce X-API-Key.
    """
    key = os.getenv("RADAR_API_KEY")
    if key and x_api_key != key:
        raise HTTPException(401, "Chybí nebo nesouhlasí X-API-Key.")


def _find_subject(db: Session, ico: str) -> Subject:
    subject = db.scalar(select(Subject).where(Subject.ico == ico))
    if not subject:
        subject = db.scalar(select(Subject).where(Subject.ico == ico.lstrip("0")))
    if not subject:
        raise HTTPException(404, f"SVJ s IČO {ico} není v databázi.")
    return subject


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

@app.get("/", response_class=HTMLResponse)
def dashboard():
    index = STATIC_DIR / "dashboard.html"
    if not index.exists():
        raise HTTPException(500, "Chybí app/static/dashboard.html")
    return index.read_text(encoding="utf-8")


# ---------------------------------------------------------------------------
# Servisní endpointy
# ---------------------------------------------------------------------------

@app.get("/api/health")
def health():
    return {"status": "ok", "service": "rbd-radar", "version": "0.3.0"}


@app.get("/api/stats")
def stats(db: Session = Depends(get_db)):
    return {
        "subjects": db.scalar(select(func.count(Subject.id))) or 0,
        "documents": db.scalar(select(func.count(Document.id))) or 0,
        "signals": db.scalar(select(func.count(Signal.id))) or 0,
        "hot_leads": db.scalar(
            select(func.count(Document.id)).where(Document.score >= 60)
        ) or 0,
    }


@app.get("/api/subjects")
def subjects(limit: int = 20, city: str | None = None,
             search: str | None = None, db: Session = Depends(get_db)):
    q = select(Subject)
    if city:
        q = q.where(Subject.city.ilike(f"%{city}%"))
    if search:
        q = q.where(Subject.name.ilike(f"%{search}%"))
    rows = db.scalars(q.order_by(Subject.name).limit(limit)).all()
    return [{
        "ico": s.ico,
        "name": s.name,
        "address": s.address,
        "city": s.city,
        "street": s.street,
        "house_number": s.house_number,
        "zip_code": s.zip_code,
        "court": s.court,
        "file_number": s.file_number,
        "last_entry_date": s.last_entry_date,
        "source_dataset": s.source_dataset,
        "listiny_checked_at": s.listiny_checked_at,
    } for s in rows]


class PrvotkarImportIn(BaseModel):
    obec: str | None = None
    okres: str | None = None
    ulice: str | None = None
    cast_obce: str | None = None
    typ: str | None = None
    limit: int | None = None


@app.post("/api/import/prvotkar", dependencies=[Depends(require_api_key)])
def prvotkar_import(payload: PrvotkarImportIn):
    """Import SVJ/BD z Prvotkáře podle obce nebo okresu."""
    from .prvotkar_client import import_obec
    if not payload.obec and not payload.okres:
        raise HTTPException(400, "Zadejte obec nebo okres.")
    try:
        out = import_obec(payload.obec, payload.ulice, payload.cast_obce,
                          payload.typ, payload.limit, okres=payload.okres)
    except Exception as exc:
        raise HTTPException(502, f"Import z Prvotkáře selhal: {exc}")
    return {"status": "ok", **out}


@app.get("/api/leads/{ico}")
def lead_by_ico(ico: str, db: Session = Depends(get_db)):
    """Lead jednoho SVJ podle IČO — pro integraci s Prvotkářem."""
    subject = _find_subject(db, ico)
    docs = db.scalars(
        select(Document).where(Document.subject_id == subject.id)
        .order_by(desc(Document.score))
    ).all()
    best = docs[0].score if docs and docs[0].score else 0
    return {
        "ico": subject.ico,
        "name": subject.name,
        "address": subject.address,
        "score": best,
        "lead_level": lead_level(best),
        "documents": [{
            "title": d.title,
            "doc_type": d.doc_type,
            "date": d.meeting_date or d.document_date,
            "score": d.score or 0,
            "source_url": d.source_url,
            "signals": [{
                "label": s.label or s.keyword,
                "value": s.value,
                "priority": s.priority,
                "keyword": s.keyword,
                "evidence": s.evidence,
            } for s in sorted(d.signals,
                              key=lambda x: x.priority or 0, reverse=True)],
        } for d in docs],
    }


@app.post("/api/import/justice", dependencies=[Depends(require_api_key)])
def justice_import(payload: JusticeImportIn):
    if not payload.dataset.startswith(("svj-", "druzstvo-", "bd-")):
        raise HTTPException(400, "Povoleny jsou pouze SVJ/druzstvo datasety.")
    import_dataset(payload.dataset, payload.limit)
    return {"status": "ok", "dataset": payload.dataset, "limit": payload.limit}


@app.post("/api/subjects", dependencies=[Depends(require_api_key)])
def create_subject(payload: SubjectIn, db: Session = Depends(get_db)):
    existing = db.scalar(select(Subject).where(Subject.ico == payload.ico))
    if existing:
        return {"id": existing.id, "created": False}
    s = Subject(**payload.model_dump())
    db.add(s)
    db.commit()
    db.refresh(s)
    return {"id": s.id, "created": True}


# ---------------------------------------------------------------------------
# Dokumenty
# ---------------------------------------------------------------------------

@app.post("/api/documents", dependencies=[Depends(require_api_key)])
def ingest_document(payload: DocumentIn, db: Session = Depends(get_db)):
    subject = _find_subject(db, payload.ico)
    result = ingest_text(
        db, subject,
        text=payload.text,
        external_id=payload.external_id,
        title=payload.title,
        source_url=payload.source_url,
        document_date=payload.document_date,
    )
    return result


@app.post("/api/documents/upload", dependencies=[Depends(require_api_key)])
async def upload_document(ico: str = Form(...),
                          file: UploadFile = File(...),
                          db: Session = Depends(get_db)):
    """Nahrání PDF listiny ručně (např. stažené z justice.cz v prohlížeči)."""
    subject = _find_subject(db, ico)
    if not (file.filename or "").lower().endswith(".pdf"):
        raise HTTPException(400, "Očekávám PDF soubor.")

    uploads = Path("data/listiny") / subject.ico
    uploads.mkdir(parents=True, exist_ok=True)
    safe_name = Path(file.filename).name
    target = uploads / safe_name
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name
    shutil.move(tmp_path, target)

    result = ingest_pdf(db, subject, target,
                        external_id=f"UPLOAD-{safe_name}",
                        title=safe_name)
    if result.get("error"):
        raise HTTPException(422, result["error"])
    return result


class SyncRunIn(BaseModel):
    limit: int = 20
    city: str | None = None
    max_docs: int = 3
    since_days: int | None = None
    icos: list[str] | None = None


@app.post("/api/sync/run", dependencies=[Depends(require_api_key)])
def sync_run(payload: SyncRunIn):
    """Spustí synchronizaci se Sbírkou listin na pozadí.

    Příklady: {"since_days": 30, "limit": 50} — nové zápisy za měsíc;
              {"city": "Brno", "limit": 20} — jen vybrané město;
              {"icos": ["3438546", …]} — konkrétní subjekty (např. okres).
    """
    if payload.icos and len(payload.icos) > 500:
        raise HTTPException(400, "Najednou lze synchronizovat max. 500 subjektů.")
    if not _start_sync(payload.limit, payload.city, payload.max_docs,
                       payload.since_days, payload.icos):
        raise HTTPException(409, "Synchronizace už běží — počkejte, až doběhne.")
    return {"status": "started",
            **{k: v for k, v in payload.model_dump().items() if k != "icos"},
            "icos_count": len(payload.icos or [])}


@app.get("/api/sync/status")
def sync_status():
    return SYNC_STATE


@app.post("/api/subjects/{ico}/sync-listiny", dependencies=[Depends(require_api_key)])
def sync_listiny(ico: str, payload: SyncIn | None = None,
                 db: Session = Depends(get_db)):
    """Stáhne a zpracuje nové listiny SVJ ze Sbírky listin justice.cz."""
    from .pipeline import sync_subject
    subject = _find_subject(db, ico)
    try:
        return sync_subject(db, subject,
                            max_docs=(payload.max_docs if payload else 5))
    except Exception as exc:
        raise HTTPException(
            502,
            f"Synchronizace se Sbírkou listin selhala: {exc}. "
            f"Web or.justice.cz omezuje frekvenci přístupů — zkuste to za chvíli."
        )


# ---------------------------------------------------------------------------
# Leady
# ---------------------------------------------------------------------------

@app.get("/api/leads")
def leads(min_score: int = 1, limit: int = 100, city: str | None = None,
          db: Session = Depends(get_db)):
    """Leady seskupené podle SVJ; skóre subjektu = nejlepší dokument."""
    q = (select(Document, Subject)
         .join(Subject, Document.subject_id == Subject.id)
         .where(Document.score >= min_score)
         .order_by(desc(Document.score), desc(Document.document_date)))
    if city:
        q = q.where(Subject.city.ilike(f"%{city}%"))
    rows = db.execute(q).all()

    by_subject: dict[int, dict] = {}
    for doc, subject in rows:
        entry = by_subject.setdefault(subject.id, {
            "ico": subject.ico,
            "name": subject.name,
            "address": subject.address,
            "city": subject.city,
            "score": 0,
            "lead_level": "LOW",
            "documents": [],
        })
        signals = db.scalars(
            select(Signal).where(Signal.document_id == doc.id)
            .order_by(desc(Signal.priority))
        ).all()
        entry["documents"].append({
            "document_id": doc.id,
            "external_id": doc.external_id,
            "title": doc.title,
            "doc_type": doc.doc_type,
            "document_date": doc.document_date,
            "meeting_date": doc.meeting_date,
            "source_url": doc.source_url,
            "score": doc.score or 0,
            "ocr_used": doc.ocr_used,
            "signals": [{
                "type": s.type,
                "label": s.label or s.keyword,
                "category": s.category,
                "priority": s.priority,
                "points": s.points,
                "value": s.value,
                "evidence": s.evidence,
                "keyword": s.keyword,
            } for s in signals],
        })
        if (doc.score or 0) > entry["score"]:
            entry["score"] = doc.score or 0
            entry["lead_level"] = lead_level(entry["score"])

    result = sorted(by_subject.values(), key=lambda x: x["score"], reverse=True)
    return result[:limit]


# ---------------------------------------------------------------------------
# Export pro obchodníky
# ---------------------------------------------------------------------------

@app.get("/api/export/leads.xlsx")
def export_leads(min_score: int = 35, city: str | None = None,
                 db: Session = Depends(get_db)):
    """Excel se seznamem leadů pro obchodníky."""
    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    data = leads(min_score=min_score, limit=1000, city=city, db=db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leady"
    headers = ["Skóre", "Úroveň", "SVJ", "IČO", "Adresa", "Město",
               "Signály", "Hodnoty", "Poslední zápis", "Odkaz na listinu"]
    ws.append(headers)
    level_fill = {"HOT": "DC2626", "HIGH": "EA580C", "WATCH": "CA8A04"}
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E293B")

    for lead in data:
        seen, sig_labels, sig_values = set(), [], []
        last_date, source = None, None
        for doc in lead["documents"]:
            d = doc.get("meeting_date") or doc.get("document_date")
            if d and (last_date is None or d > last_date):
                last_date, source = d, doc.get("source_url")
            for s in doc["signals"]:
                if s["label"] in seen:
                    continue
                seen.add(s["label"])
                sig_labels.append(s["label"])
                if s.get("value"):
                    sig_values.append(f"{s['label']}: {s['value']}")
        ws.append([
            lead["score"], lead["lead_level"], lead["name"], lead["ico"],
            lead.get("address") or "", lead.get("city") or "",
            ", ".join(sig_labels), ", ".join(sig_values),
            last_date.strftime("%d.%m.%Y") if last_date else "",
            source or "",
        ])
        fill = level_fill.get(lead["lead_level"])
        if fill:
            ws.cell(row=ws.max_row, column=2).fill = PatternFill(
                "solid", fgColor=fill)
            ws.cell(row=ws.max_row, column=2).font = Font(
                bold=True, color="FFFFFF")

    widths = [8, 9, 46, 11, 34, 16, 44, 40, 14, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"rbd-radar-leady{('-' + city) if city else ''}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument"
                   ".spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Demo
# ---------------------------------------------------------------------------

@app.post("/api/demo", dependencies=[Depends(require_api_key)])
def demo(db: Session = Depends(get_db)):
    ico = "12345678"
    subject = db.scalar(select(Subject).where(Subject.ico == ico))
    if not subject:
        subject = Subject(
            ico=ico,
            name="SVJ Demo – Jihlava",
            legal_form="Společenství vlastníků jednotek",
            address="Jihlava",
            source_url="https://or.justice.cz/ias/ui/rejstrik",
        )
        db.add(subject)
        db.commit()
        db.refresh(subject)
    text = (
        "Zápis ze shromáždění vlastníků jednotek konané dne 12.5.2026. "
        "Shromáždění projednalo přípravu komplexní revitalizace bytového domu. "
        "Byla předložena nabídka na zateplení fasády, rekonstrukci balkonů "
        "a hydroizolaci střechy. Výbor byl pověřen zajištěním projektové "
        "dokumentace a prověřením financování prostřednictvím NZÚ a úvěru. "
        "Bylo schváleno navýšení příspěvku do fondu oprav na 45 Kč/m2."
    )
    return ingest_text(
        db, subject,
        text=text,
        external_id=f"DEMO-{hashlib.sha256(text.encode()).hexdigest()[:8]}",
        title="Zápis ze shromáždění – příprava revitalizace",
        source_url="https://or.justice.cz/ias/ui/rejstrik",
    )
